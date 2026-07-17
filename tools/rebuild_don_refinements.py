#!/usr/bin/env python3
"""Rebuild DoN BiS block with Dragons of Norrath refinements (v1.2.18+)."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(r"d:\Desktop\Turbo Wednesday 17th\DoN Misc..xlsx")
OUT_BLOCK = Path(__file__).resolve().parent / "don_bis_block.lua"
OUT_ANNOUNCE = Path(__file__).resolve().parent / "don_announce_names.txt"
BIS_LUA = Path(r"D:\Desktop\Turbo v3.9.88\Turbo v3.9.88 git\lazbis\bis.lua")
ANNOUNCE_INI = Path(
    r"D:\Desktop\Turbo v3.9.88\Turbo v3.9.88 git\Turbo v3.9.88\lua\Turbo\rulepacks\BiS_announce_list.ini"
)

ABBREV_TO_CLASS = {
    "WAR": "Warrior",
    "SHD": "Shadow Knight",
    "PAL": "Paladin",
    "CLR": "Cleric",
    "SHM": "Shaman",
    "DRU": "Druid",
    "MNK": "Monk",
    "ROG": "Rogue",
    "BER": "Berserker",
    "BRD": "Bard",
    "RNG": "Ranger",
    "BST": "Beastlord",
    "WIZ": "Wizard",
    "MAG": "Magician",
    "ENC": "Enchanter",
    "NEC": "Necromancer",
}

GLYPH_CLASS_ALIASES = {
    "Cleric": "Cleric",
    "Shaman": "Shaman",
    "Druid": "Druid",
    "Wizard": "Wizard",
    "Magician": "Magician",
    "Enchanter": "Enchanter",
    "Necromancer": "Necromancer",
    "Shadowknight": "Shadow Knight",
    "Shadow Knight": "Shadow Knight",
    "Paladin": "Paladin",
    "Ranger": "Ranger",
    "Beastlord": "Beastlord",
    "Bard": "Bard",
    "Warrior": "Warrior",
    "Monk": "Monk",
    "Rogue": "Rogue",
    "Berserker": "Berserker",
}

# First = Ranged, second = RangedAug
RANGED_BY_CLASS = {
    "Beastlord": (("Branch of the Twisting Tree", 71627), ("Fallen Leaf of the Twisting Tree", 55053)),
    "Shadow Knight": (("Branch of the Twisting Tree", 71627), ("Fallen Leaf of the Twisting Tree", 55053)),
    "Ranger": (("Branch of the Twisting Tree", 71627), ("Fallen Leaf of the Twisting Tree", 55053)),
    "Paladin": (("Branch of the Twisting Tree", 71627), ("Fallen Leaf of the Twisting Tree", 55053)),
    "Berserker": (("Head of the Putrid Drake", 55052), ("Preserved Eye of the Putrid Drake", 71620)),
    "Rogue": (("Head of the Putrid Drake", 55052), ("Preserved Eye of the Putrid Drake", 71620)),
    "Bard": (("Head of the Putrid Drake", 55052), ("Preserved Eye of the Putrid Drake", 71620)),
    "Monk": (("Head of the Putrid Drake", 55052), ("Preserved Eye of the Putrid Drake", 71620)),
    "Warrior": (("Head of the Putrid Drake", 55052), ("Preserved Eye of the Putrid Drake", 71620)),
    "Enchanter": (("Spark of the Skies", 60536), ("Core of the Skies", 71659)),
    "Magician": (("Spark of the Skies", 60536), ("Core of the Skies", 71659)),
    "Wizard": (("Spark of the Skies", 60536), ("Core of the Skies", 71659)),
    "Necromancer": (("Spark of the Skies", 60536), ("Core of the Skies", 71659)),
    "Shaman": (("Crux of the First Brood", 60535), ("Phial of the First Brood", 71578)),
    "Druid": (("Crux of the First Brood", 60535), ("Phial of the First Brood", 71578)),
    "Cleric": (("Crux of the First Brood", 60535), ("Phial of the First Brood", 71578)),
}

ICON_BY_CLASS = {
    "Shadow Knight": ("Icon of Unwavering Defense", 56755),
    "Paladin": ("Icon of Unwavering Defense", 56755),
    "Warrior": ("Icon of Unwavering Defense", 56755),
    "Enchanter": ("Icon of Scribe's Endurance", 56756),
    "Magician": ("Icon of Scribe's Endurance", 56756),
    "Wizard": ("Icon of Scribe's Endurance", 56756),
    "Necromancer": ("Icon of Scribe's Endurance", 56756),
    "Shaman": ("Icon of Scribe's Endurance", 56756),
    "Druid": ("Icon of Scribe's Endurance", 56756),
    "Cleric": ("Icon of Scribe's Endurance", 56756),
    "Berserker": ("Icon of Potent Prowess", 56757),
    "Beastlord": ("Icon of Potent Prowess", 56757),
    "Rogue": ("Icon of Potent Prowess", 56757),
    "Bard": ("Icon of Potent Prowess", 56757),
    "Monk": ("Icon of Potent Prowess", 56757),
    "Ranger": ("Icon of Potent Prowess", 56757),
}

# Short display | full aug | vessel | effect | bonus (for colored tooltip notes)
FOCI_META = {
    "Malevolent Efficiency": {
        "aug": 62540,
        "vessel": 62560,
        "effect": "-11-33% detrimental spell mana cost",
        "bonus": "+10 hINT",
        "tags": ["int"],
    },
    "Benevolent Efficiency": {
        "aug": 62541,
        "vessel": 62561,
        "effect": "-11-33% beneficial spell mana cost",
        "bonus": "+10 hWIS",
        "tags": ["wis"],
    },
    "Malevolent Alacrity": {
        "aug": 62542,
        "vessel": 62562,
        "effect": "-40% detrimental spell cast time",
        "bonus": "+10 hINT",
        "tags": ["int"],
    },
    "Benevolent Alacrity": {
        "aug": 62543,
        "vessel": 62563,
        "effect": "-40% beneficial spell cast time",
        "bonus": "+10 hWIS",
        "tags": ["wis"],
    },
    "Arcane Demise": {
        "aug": 62544,
        "vessel": 62564,
        "effect": "+10-60% magic spell damage",
        "bonus": "+5 hINT, +25 hMR",
        "tags": ["magic", "int"],
    },
    "Fiery Demise": {
        "aug": 62545,
        "vessel": 62565,
        "effect": "+10-60% fire spell damage",
        "bonus": "+5 hINT, +25 hFR",
        "tags": ["fire", "int"],
    },
    "Chilling Demise": {
        "aug": 62546,
        "vessel": 62566,
        "effect": "+10-60% cold spell damage",
        "bonus": "+5 hINT, +25 hCR",
        "tags": ["cold", "int"],
    },
    "Noxious Demise": {
        "aug": 62547,
        "vessel": 62567,
        "effect": "+10-60% poison spell damage",
        "bonus": "+5 hINT, +25 hPR",
        "tags": ["poison", "int"],
    },
    "Festering Demise": {
        "aug": 62548,
        "vessel": 62568,
        "effect": "+10-60% disease spell damage",
        "bonus": "+5 hINT, +25 hDR",
        "tags": ["disease", "int"],
    },
    "Merciful Mending": {
        "aug": 62549,
        "vessel": 62569,
        "effect": "+10-60% healing",
        "bonus": "+10 hWIS",
        "tags": ["heal", "wis"],
    },
    "Malevolent Extension": {
        "aug": 62550,
        "vessel": 63016,
        "effect": "+40% detrimental spell duration",
        "bonus": "+10 hINT",
        "tags": ["int"],
    },
    "Benevolent Extension": {
        "aug": 62551,
        "vessel": 63017,
        "effect": "+40% beneficial spell duration",
        "bonus": "+10 hWIS",
        "tags": ["wis"],
    },
    "Expanded Reach": {
        "aug": 62552,
        "vessel": 63018,
        "effect": "+50% spell range",
        "bonus": "+10 hWIS, +10 hINT",
        "tags": ["wis", "int"],
    },
    "Wanton Assault": {
        "aug": 62553,
        "vessel": 63019,
        "effect": "+21% double attack, +7% triple attack, +5% chance to hit",
        "bonus": "+10 hSTR",
        "tags": ["str"],
    },
    "Visceral Malice": {
        "aug": 62554,
        "vessel": 63020,
        "effect": "+300% melee critical damage",
        "bonus": "+10 hSTR",
        "tags": ["str"],
    },
    "Adept Guard": {
        "aug": 62555,
        "vessel": 64049,
        "effect": "+70% parry, +70% block",
        "bonus": "+10 hSTA",
        "tags": ["sta"],
    },
    "Nimble Elusion": {
        "aug": 62556,
        "vessel": 64050,
        "effect": "+70% dodge",
        "bonus": "+10 hAGI",
        "tags": ["agi"],
    },
    "Lethal Barrage": {
        "aug": 62557,
        "vessel": 64326,
        "effect": "+28% archery and throwing chance to hit",
        "bonus": "+20 hDEX",
        "tags": ["dex"],
    },
    "Physical Prowess": {
        "aug": 62558,
        "vessel": 66533,
        "effect": "No focus effect",
        "bonus": "+5 hSTR, +5 hSTA, +5 hAGI, +5 hDEX",
        "tags": ["str", "sta", "agi", "dex"],
    },
    "Mental Prowess": {
        "aug": 62559,
        "vessel": 66720,
        "effect": "No focus effect",
        "bonus": "+10 hINT, +10 hWIS, +10 heroic resists",
        "tags": ["int", "wis"],
    },
}

SLOT_MAP = {
    "Head": "Head",
    "Chest": "Chest",
    "Arms": "Arms",
    "Wrist 1": "Wrist1",
    "Wrist1": "Wrist1",
    "Hands": "Hands",
    "Legs": "Legs",
    "Feet": "Feet",
    "Wrist 2": "Wrist2",
    "Wrist2": "Wrist2",
}

CLASS_ORDER = [
    "Bard",
    "Beastlord",
    "Berserker",
    "Cleric",
    "Druid",
    "Enchanter",
    "Magician",
    "Monk",
    "Necromancer",
    "Paladin",
    "Ranger",
    "Rogue",
    "Shadow Knight",
    "Shaman",
    "Warrior",
    "Wizard",
]

# Copied from prior DoN plan (DSK Hideous Hex → Cryptic Clutch; skip Companion)
FOCI_BY_CLASS = {
    "Warrior": [
        "Benevolent Extension", "Fiery Demise", "Noxious Demise", "Nimble Elusion",
        "Adept Guard", "Visceral Malice", "Wanton Assault", "Physical Prowess",
    ],
    "Shadow Knight": [
        "Benevolent Extension", "Malevolent Extension", "Noxious Demise", "Arcane Demise",
        "Festering Demise", "Expanded Reach", "Nimble Elusion", "Adept Guard",
        "Visceral Malice", "Wanton Assault", "Physical Prowess",
    ],
    "Paladin": [
        "Benevolent Efficiency", "Benevolent Extension", "Arcane Demise", "Expanded Reach",
        "Merciful Mending", "Nimble Elusion", "Adept Guard", "Visceral Malice",
        "Wanton Assault", "Physical Prowess",
    ],
    "Cleric": [
        "Benevolent Efficiency", "Benevolent Extension", "Benevolent Alacrity",
        "Malevolent Efficiency", "Malevolent Extension", "Malevolent Alacrity",
        "Arcane Demise", "Expanded Reach", "Merciful Mending", "Nimble Elusion", "Mental Prowess",
    ],
    "Shaman": [
        "Benevolent Efficiency", "Benevolent Extension", "Benevolent Alacrity",
        "Malevolent Efficiency", "Malevolent Extension", "Malevolent Alacrity",
        "Noxious Demise", "Festering Demise", "Expanded Reach", "Merciful Mending",
        "Nimble Elusion", "Mental Prowess",
    ],
    "Druid": [
        "Benevolent Efficiency", "Benevolent Extension", "Benevolent Alacrity",
        "Malevolent Efficiency", "Malevolent Extension", "Malevolent Alacrity",
        "Fiery Demise", "Arcane Demise", "Expanded Reach", "Merciful Mending",
        "Nimble Elusion", "Mental Prowess",
    ],
    "Monk": [
        "Benevolent Extension", "Noxious Demise", "Nimble Elusion", "Adept Guard",
        "Visceral Malice", "Wanton Assault", "Physical Prowess",
    ],
    "Rogue": [
        "Benevolent Extension", "Malevolent Extension", "Noxious Demise", "Nimble Elusion",
        "Adept Guard", "Visceral Malice", "Wanton Assault", "Physical Prowess",
    ],
    "Berserker": [
        "Benevolent Extension", "Malevolent Extension", "Noxious Demise", "Nimble Elusion",
        "Adept Guard", "Visceral Malice", "Wanton Assault", "Lethal Barrage", "Physical Prowess",
    ],
    "Bard": [
        "Benevolent Extension", "Malevolent Efficiency", "Expanded Reach", "Nimble Elusion",
        "Adept Guard", "Visceral Malice", "Wanton Assault", "Physical Prowess",
    ],
    "Ranger": [
        "Benevolent Extension", "Malevolent Efficiency", "Malevolent Extension", "Fiery Demise",
        "Arcane Demise", "Expanded Reach", "Merciful Mending", "Nimble Elusion", "Adept Guard",
        "Visceral Malice", "Wanton Assault", "Lethal Barrage",
    ],
    "Beastlord": [
        "Benevolent Extension", "Malevolent Efficiency", "Malevolent Extension", "Noxious Demise",
        "Festering Demise", "Merciful Mending", "Nimble Elusion", "Adept Guard", "Visceral Malice",
        "Wanton Assault", "Chilling Demise", "Physical Prowess",
    ],
    "Wizard": [
        "Benevolent Extension", "Malevolent Efficiency", "Malevolent Extension", "Malevolent Alacrity",
        "Fiery Demise", "Arcane Demise", "Expanded Reach", "Nimble Elusion", "Mental Prowess",
    ],
    "Magician": [
        "Benevolent Extension", "Benevolent Alacrity", "Malevolent Efficiency", "Malevolent Extension",
        "Malevolent Alacrity", "Fiery Demise", "Expanded Reach", "Nimble Elusion", "Mental Prowess",
    ],
    "Enchanter": [
        "Benevolent Efficiency", "Benevolent Extension", "Benevolent Alacrity",
        "Malevolent Efficiency", "Malevolent Extension", "Malevolent Alacrity",
        "Arcane Demise", "Expanded Reach", "Nimble Elusion", "Mental Prowess",
    ],
    "Necromancer": [
        "Benevolent Extension", "Benevolent Alacrity", "Malevolent Efficiency", "Malevolent Extension",
        "Malevolent Alacrity", "Fiery Demise", "Noxious Demise", "Arcane Demise", "Festering Demise",
        "Expanded Reach", "Nimble Elusion", "Mental Prowess",
    ],
}


def lua_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def item_line(name: str, iid) -> str:
    name = str(name).strip()
    if iid is None or iid == "" or iid == "-":
        return name
    return f"{name}/{int(iid)}"


def parse_classes(cell) -> list[str]:
    if cell is None:
        return []
    text = str(cell).strip()
    if not text:
        return []
    upper = text.upper()
    if upper == "ALL":
        return list(ABBREV_TO_CLASS.values())
    out = []
    for part in re.split(r"[/|,]+", text):
        part = part.strip().upper()
        if part in ABBREV_TO_CLASS:
            out.append(ABBREV_TO_CLASS[part])
    return out


def emit_value(val, indent: str) -> str:
    if isinstance(val, dict):
        item = val["item"]
        parts = [f"{{\n{indent}\titem = '{lua_escape(item)}',"]
        if val.get("spells"):
            spell_lits = ", ".join(f"'{lua_escape(s)}'" for s in val["spells"])
            parts.append(f"\n{indent}\tspells = {{{spell_lits}}},")
        if val.get("names"):
            name_lits = ", ".join(f"'{lua_escape(n)}'" for n in val["names"])
            parts.append(f"\n{indent}\tnames = {{{name_lits}}},")
        if val.get("ids"):
            id_lits = ", ".join(str(i) for i in val["ids"])
            parts.append(f"\n{indent}\tids = {{{id_lits}}},")
        if val.get("notes"):
            parts.append(f"\n{indent}\tnotes = '{lua_escape(val['notes'])}',")
        parts.append(f"\n{indent}}}")
        return "".join(parts)
    return f"'{lua_escape(val)}'"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    classes: dict[str, dict] = {c: {} for c in CLASS_ORDER}
    announce: set[str] = set()

    def collect(name: str):
        if name:
            announce.add(str(name).split("/")[0].strip())

    # Visibles
    ws = wb["Visible Icons"]
    for r in range(2, ws.max_row + 1):
        abbr = ws.cell(r, 2).value
        slot_raw = ws.cell(r, 3).value
        iid = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if not abbr or not name:
            continue
        cls = ABBREV_TO_CLASS.get(str(abbr).strip().upper())
        slot = SLOT_MAP.get(str(slot_raw).strip())
        if not cls or not slot:
            continue
        classes[cls][slot] = item_line(name, iid)
        collect(name)

    # Non-visibles (skip clickies/range/charm/epic/legendary material)
    ear_lists: dict[str, list] = defaultdict(list)
    ring_lists: dict[str, list] = defaultdict(list)
    ws = wb["Non-Visible"]
    for r in range(2, ws.max_row + 1):
        slot_raw = ws.cell(r, 1).value
        iid = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        classes_cell = ws.cell(r, 5).value
        if not slot_raw or not name:
            continue
        slot_raw = str(slot_raw).strip().strip('"')
        if slot_raw in ("Range", "Charm", "Epic", "Clicky") or "Legendary" in slot_raw:
            continue
        if str(name).startswith("Primary/"):
            continue
        targets = parse_classes(classes_cell)
        if not targets:
            continue
        line = item_line(name, iid)
        collect(name)
        if slot_raw == "Ear":
            for c in targets:
                ear_lists[c].append(line)
        elif slot_raw == "Ring":
            for c in targets:
                ring_lists[c].append(line)
        else:
            slot = {
                "Face": "Face",
                "Neck": "Neck",
                "Back": "Back",
                "Shoulder": "Shoulder",
                "Waist": "Waist",
            }.get(slot_raw)
            if not slot:
                continue
            for c in targets:
                classes[c][slot] = line

    for c, ears in ear_lists.items():
        if ears:
            classes[c]["Ear1"] = ears[0]
        if len(ears) >= 2:
            classes[c]["Ear2"] = ears[1]
    for c, rings in ring_lists.items():
        if rings:
            classes[c]["Finger1"] = rings[0]
        if len(rings) >= 2:
            classes[c]["Finger2"] = rings[1]

    # Ranged + RangedAug
    for c, (ranged, aug) in RANGED_BY_CLASS.items():
        classes[c]["Ranged"] = item_line(*ranged)
        classes[c]["RangedAug"] = item_line(*aug)
        collect(ranged[0])
        collect(aug[0])

    # Icons (single Clicky)
    for c, (name, iid) in ICON_BY_CLASS.items():
        classes[c]["Clicky"] = item_line(name, iid)
        collect(name)

    # Legendary → Shadow
    ws = wb["Class WeaponShield Components"]
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, 1).value
        name = ws.cell(r, 3).value
        abbr = ws.cell(r, 4).value
        if not name or not abbr or str(abbr).upper() == "ALL":
            continue
        if "Materium" in str(name):
            continue
        cls = ABBREV_TO_CLASS.get(str(abbr).strip().upper())
        if cls:
            classes[cls]["Shadow"] = item_line(name, iid)
            collect(name)

    # Glyphs
    glyph_by_template: dict[int, tuple[str, str]] = {}
    glyphs_by_class: dict[str, list] = defaultdict(list)
    ws = wb["Spell Clicky Components"]
    for r in range(2, ws.max_row + 1):
        gclass = ws.cell(r, 1).value
        spell = ws.cell(r, 2).value
        tmpl_id = ws.cell(r, 4).value
        clicky_name = ws.cell(r, 6).value
        clicky_id = ws.cell(r, 7).value
        if not gclass or not clicky_name:
            continue
        cls = GLYPH_CLASS_ALIASES.get(str(gclass).strip())
        if not cls:
            continue
        glyphs_by_class[cls].append(item_line(clicky_name, clicky_id))
        collect(clicky_name)
        if tmpl_id:
            glyph_by_template[int(tmpl_id)] = (cls, str(spell))
    for c, glyphs in glyphs_by_class.items():
        for i, line in enumerate(glyphs, start=1):
            classes[c][f"Glyph{i}"] = line

    # Packs
    pack_counts: dict[str, int] = defaultdict(int)
    ws = wb["Spell Packs"]
    for r in range(2, ws.max_row + 1):
        abbr = ws.cell(r, 1).value
        iid = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        if not abbr or not name:
            continue
        cls = ABBREV_TO_CLASS.get(str(abbr).strip().upper())
        if not cls:
            continue
        tmpl_ids = []
        for col in (7, 9, 11):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                tmpl_ids.append(int(v))
        spells = []
        for tid in tmpl_ids:
            if tid in glyph_by_template:
                spells.append(glyph_by_template[tid][1])
        if not spells:
            m = re.match(r"^(?:Spell|Tome) Pack:\s*(.+)$", str(name).strip())
            if m:
                spells = [m.group(1).strip()]
        pack_counts[cls] += 1
        key = f"Pack{pack_counts[cls]}"
        line = item_line(name, iid)
        collect(name)
        if spells:
            classes[cls][key] = {"item": line, "spells": spells}
        else:
            classes[cls][key] = line

    # Foci - short display, match aug + vessel
    for c, shorts in FOCI_BY_CLASS.items():
        for i, short in enumerate(shorts, start=1):
            meta = FOCI_META[short]
            full = f"Cryptic Clutch of {short}"
            vessel = f"Vacant Vessel of {short}"
            notes = f"{meta['effect']} | {meta['bonus']}"
            classes[c][f"Aug{i}"] = {
                "item": short,
                "names": [short, full, vessel],
                "ids": [meta["aug"], meta["vessel"]],
                "notes": notes,
            }
            collect(full)
            collect(vessel)

    # Template shared
    template = {
        "CharmExtreme": "Lustrous Gem of Eternal Avarice/55054",
        "CharmSafe": "Solemn Gem of Restrained Avarice/55055",
        "Duality": "Duality of Desire/66903",
        "Materium1": "Primary Materium of Legends/62535",
        "Materium2": "Secondary Materium of Legends/62536",
        "Materium3": "Tertiary Materium of Legends/62537",
        "Misc1": "Dark Reign Elite Satchel/66731",
        "Misc2": "Dark Reign Initiate Satchel/66732",
        "Misc3": "Ancient Draconic Lockbox I/66733",
        "Misc4": "Scales of the Lava Dragon/66904",
        "Misc5": "Draconium Surveyor's Waystone/43224",
    }
    for v in template.values():
        collect(v)

    max_pack = max(pack_counts.values()) if pack_counts else 1
    max_glyph = max((len(glyphs_by_class[c]) for c in CLASS_ORDER), default=1)
    max_aug = max(len(FOCI_BY_CLASS[c]) for c in CLASS_ORDER)
    pack_slots = [f"Pack{i}" for i in range(1, max_pack + 1)]
    glyph_slots = [f"Glyph{i}" for i in range(1, max_glyph + 1)]
    aug_slots = [f"Aug{i}" for i in range(1, max_aug + 1)]

    lines = ["\t['don'] = {"]
    for c in CLASS_ORDER:
        bucket = classes[c]
        lines.append(f"\t\t['{c}'] = {{")
        preferred = (
            ["Head", "Chest", "Arms", "Wrist1", "Hands", "Legs", "Feet", "Wrist2"]
            + ["Ear1", "Ear2", "Face", "Neck", "Back", "Shoulder", "Waist", "Finger1", "Finger2"]
            + ["Ranged", "RangedAug", "Clicky", "Shadow"]
            + pack_slots
            + glyph_slots
            + aug_slots
        )
        seen = set()
        for slot in preferred:
            if slot in bucket:
                lines.append(f"\t\t\t['{slot}'] = {emit_value(bucket[slot], chr(9)*3)},")
                seen.add(slot)
        for slot in sorted(bucket.keys()):
            if slot not in seen:
                lines.append(f"\t\t\t['{slot}'] = {emit_value(bucket[slot], chr(9)*3)},")
        lines.append("\t\t},")

    lines.append("\t\t['Main'] = {")
    lines.append("\t\t\t['Slots'] = {")
    lines.append(
        "\t\t\t\t{Name='Visibles', Slots={'Head','Chest','Arms','Wrist1','Hands','Legs','Feet','Wrist2',}},"
    )
    lines.append(
        "\t\t\t\t{Name='Non-Visibles', Slots={'Ear1','Ear2','Face','Neck','Back','Shoulder','Waist','Finger1','Finger2',}},"
    )
    lines.append(
        "\t\t\t\t{Name='Ranged + Charm', Slots={'Ranged','RangedAug','Duality','CharmExtreme','CharmSafe',}},"
    )
    pack_lua = ",".join(f"'{s}'" for s in pack_slots)
    lines.append(f"\t\t\t\t{{Name='Spells', Slots={{{pack_lua},}}}},")
    glyph_lua = ",".join(f"'{s}'" for s in glyph_slots)
    lines.append(f"\t\t\t\t{{Name='Glyphs', Slots={{{glyph_lua},}}}},")
    aug_lua = ",".join(f"'{s}'" for s in aug_slots)
    lines.append(f"\t\t\t\t{{Name='Cryptic Clutch Foci Augs', Slots={{{aug_lua},}}}},")
    lines.append("\t\t\t\t{Name='Clickies', Slots={'Clicky',}},")
    lines.append(
        "\t\t\t\t{Name='Shadow', Slots={'Materium1','Materium2','Materium3','Shadow',}},"
    )
    lines.append(
        "\t\t\t\t{Name='Misc', Slots={'Misc1','Misc2','Misc3','Misc4','Misc5',}},"
    )
    lines.append("\t\t\t},")
    lines.append("\t\t},")
    lines.append("\t\t['Template'] = {")
    for k, v in template.items():
        lines.append(f"\t\t\t['{k}'] = '{lua_escape(v)}',")
    lines.append("\t\t},")
    lines.append("\t},")

    block = "\n".join(lines) + "\n"
    OUT_BLOCK.write_text(block, encoding="utf-8")

    # Patch lazbis/bis.lua
    text = BIS_LUA.read_text(encoding="utf-8")
    text = text.replace(
        "[7]={id='don',name='DoN'},",
        "[7]={id='don',name='Dragons of Norrath'},",
    )
    new_text, n = re.subn(
        r"\n\t\['don'\] = \{.*?\n\t\},\n\t\['jonas'\]",
        "\n" + block.rstrip() + "\n\t['jonas']",
        text,
        count=1,
        flags=re.S,
    )
    if n != 1:
        raise SystemExit(f"don replace failed n={n}")
    BIS_LUA.write_text(new_text, encoding="utf-8")
    print(f"Patched {BIS_LUA}")

    # Announce names - drop removed items, add new
    remove_announce = {
        "Radiant Crystal Cache",
        "Guide of the Dark Reign",
        "A Strange Compass",
        "a damaged compass",
        "Faded Gem of Restrained Avarice",
        "Lustrous Gem of Unyielding Avarice",
        "Thorn of the Twisting Tree",
        "Codex of Unwavering Defense upgrade",
        "Codex of Scribe's Endurance upgrade",
        "Codex of Potent Prowess upgrade",
        "Codex of Minion's Materiel upgrade",
    }
    for n in remove_announce:
        announce.discard(n)

    sorted_names = sorted(announce, key=lambda s: s.lower())
    OUT_ANNOUNCE.write_text("\n".join(sorted_names) + "\n", encoding="utf-8")

    ini = ANNOUNCE_INI.read_text(encoding="utf-8")
    if ";DoN" in ini:
        head, _, _ = ini.partition(";DoN")
        # keep everything before ;DoN
        ini = head.rstrip("\n") + "\n"
    block_ini = ["", ";DoN / Dragons of Norrath"] + [f"{n}=ANNOUNCE" for n in sorted_names] + [""]
    ANNOUNCE_INI.write_text(ini.rstrip("\n") + "\n" + "\n".join(block_ini), encoding="utf-8")
    print(f"Announce names: {len(sorted_names)}")
    print("Done")


if __name__ == "__main__":
    main()
