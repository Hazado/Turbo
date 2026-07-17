import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook(r"d:\Desktop\Turbo Wednesday 17th\DoN Misc..xlsx", data_only=True)

print("=== ALL SPELL PACKS ===")
ws = wb["Spell Packs"]
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 12)]
    print(vals)

print("\n=== ALL GLYPHS ===")
ws = wb["Spell Clicky Components"]
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 8)]
    print(vals)

print("\n=== NON-VISIBLE FULL ===")
ws = wb["Non-Visible"]
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 12)]
    print(vals)

print("\n=== FOCI ===")
ws = wb["Slot 12 Foci Augs"]
for r in range(2, ws.max_row + 1):
    print(ws.cell(r, 1).value, int(ws.cell(r, 2).value) if ws.cell(r, 2).value else None)

print("\n=== MISC ===")
ws = wb["Misc. Items"]
for r in range(2, ws.max_row + 1):
    print(ws.cell(r, 1).value, ws.cell(r, 2).value)
