#!/usr/bin/env python3
"""Import DoN Misc.xlsx → lazbis ['don'] Lua block + announce name list."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(r"d:\Desktop\Turbo Wednesday 17th\DoN Misc..xlsx")
OUT_BLOCK = Path(__file__).resolve().parent / "don_bis_block.lua"
OUT_ANNOUNCE = Path(__file__).resolve().parent / "don_announce_names.txt"
OUT_SPELLS = Path(__file__).resolve().parent / "don_pack_spells.lua"

ABBREV_TO_CLASS = {
    "WAR": "Warrior",
    "SHD": "Shadow Knight",
    "PAL": "Paladin",
    "CLR": "Cleric",
    "SHM": "Shaman",
    "DRU": "Druid",
    "MNK": "Monk",
    "ROG": "Rogue",
    "BER": "Berserker",
    "BRD": "Bard",
    "RNG": "Ranger",
    "BST": "Beastlord",
    "WIZ": "Wizard",
    "MAG": "Magician",
    "ENC": "Enchanter",
    "NEC": "Necromancer",
}

CLASS_TO_ABBREV = {v: k for k, v in ABBREV_TO_CLASS.items()}

# Full class names as they appear on Spell Clicky Components sheet
GLYPH_CLASS_ALIASES = {
    "Cleric": "Cleric",
    "Shaman": "Shaman",
    "Druid": "Druid",
    "Wizard": "Wizard",
    "Magician": "Magician",
    "Enchanter": "Enchanter",
    "Necromancer": "Necromancer",
    "Shadowknight": "Shadow Knight",
    "Shadow Knight": "Shadow Knight",
    "Paladin": "Paladin",
    "Ranger": "Ranger",
    "Beastlord": "Beastlord",
    "Bard": "Bard",
    "Warrior": "Warrior",
    "Monk": "Monk",
    "Rogue": "Rogue",
    "Berserker": "Berserker",
}

RANGE_BY_CLASS = {
    "Warrior": ("Head of the Putrid Drake", 71620),
    "Monk": ("Head of the Putrid Drake", 71620),
    "Rogue": ("Head of the Putrid Drake", 71620),
    "Berserker": ("Head of the Putrid Drake", 71620),
    "Shadow Knight": ("Thorn of the Twisting Tree", 71627),
    "Paladin": ("Thorn of the Twisting Tree", 71627),
    "Bard": ("Thorn of the Twisting Tree", 71627),
    "Ranger": ("Thorn of the Twisting Tree", 71627),
    "Beastlord": ("Thorn of the Twisting Tree", 71627),
    "Wizard": ("Core of the Skies", 71659),
    "Magician": ("Core of the Skies", 71659),
    "Enchanter": ("Core of the Skies", 71659),
    "Necromancer": ("Core of the Skies", 71659),
    "Cleric": ("Phial of the First Brood", 71578),
    "Shaman": ("Phial of the First Brood", 71578),
    "Druid": ("Phial of the First Brood", 71578),
}

# DSK Hideous Hex → Cryptic Clutch (skip Companion)
FOCI_BY_CLASS = {
    "Warrior": [
        "Benevolent Extension",
        "Fiery Demise",
        "Noxious Demise",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Physical Prowess",
    ],
    "Shadow Knight": [
        "Benevolent Extension",
        "Malevolent Extension",
        "Noxious Demise",
        "Arcane Demise",
        "Festering Demise",
        "Expanded Reach",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Physical Prowess",
    ],
    "Paladin": [
        "Benevolent Efficiency",
        "Benevolent Extension",
        "Arcane Demise",
        "Expanded Reach",
        "Merciful Mending",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Physical Prowess",
    ],
    "Cleric": [
        "Benevolent Efficiency",
        "Benevolent Extension",
        "Benevolent Alacrity",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Arcane Demise",
        "Expanded Reach",
        "Merciful Mending",
        "Nimble Elusion",
        "Mental Prowess",
    ],
    "Shaman": [
        "Benevolent Efficiency",
        "Benevolent Extension",
        "Benevolent Alacrity",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Noxious Demise",
        "Festering Demise",
        "Expanded Reach",
        "Merciful Mending",
        "Nimble Elusion",
        "Mental Prowess",
    ],
    "Druid": [
        "Benevolent Efficiency",
        "Benevolent Extension",
        "Benevolent Alacrity",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Fiery Demise",
        "Arcane Demise",
        "Expanded Reach",
        "Merciful Mending",
        "Nimble Elusion",
        "Mental Prowess",
    ],
    "Monk": [
        "Benevolent Extension",
        "Noxious Demise",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Physical Prowess",
    ],
    "Rogue": [
        "Benevolent Extension",
        "Malevolent Extension",
        "Noxious Demise",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Physical Prowess",
    ],
    "Berserker": [
        "Benevolent Extension",
        "Malevolent Extension",
        "Noxious Demise",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Lethal Barrage",
        "Physical Prowess",
    ],
    "Bard": [
        "Benevolent Extension",
        "Malevolent Efficiency",
        "Expanded Reach",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Physical Prowess",
    ],
    "Ranger": [
        "Benevolent Extension",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Fiery Demise",
        "Arcane Demise",
        "Expanded Reach",
        "Merciful Mending",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Lethal Barrage",
    ],
    "Beastlord": [
        "Benevolent Extension",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Noxious Demise",
        "Festering Demise",
        "Merciful Mending",
        "Nimble Elusion",
        "Adept Guard",
        "Visceral Malice",
        "Wanton Assault",
        "Chilling Demise",
        "Physical Prowess",
    ],
    "Wizard": [
        "Benevolent Extension",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Fiery Demise",
        "Arcane Demise",
        "Expanded Reach",
        "Nimble Elusion",
        "Mental Prowess",
    ],
    "Magician": [
        "Benevolent Extension",
        "Benevolent Alacrity",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Fiery Demise",
        "Expanded Reach",
        "Nimble Elusion",
        "Mental Prowess",
    ],
    "Enchanter": [
        "Benevolent Efficiency",
        "Benevolent Extension",
        "Benevolent Alacrity",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Arcane Demise",
        "Expanded Reach",
        "Nimble Elusion",
        "Mental Prowess",
    ],
    "Necromancer": [
        "Benevolent Extension",
        "Benevolent Alacrity",
        "Malevolent Efficiency",
        "Malevolent Extension",
        "Malevolent Alacrity",
        "Fiery Demise",
        "Noxious Demise",
        "Arcane Demise",
        "Festering Demise",
        "Expanded Reach",
        "Nimble Elusion",
        "Mental Prowess",
    ],
}

SLOT_MAP = {
    "Head": "Head",
    "Chest": "Chest",
    "Arms": "Arms",
    "Wrist 1": "Wrist1",
    "Wrist1": "Wrist1",
    "Hands": "Hands",
    "Legs": "Legs",
    "Feet": "Feet",
    "Wrist 2": "Wrist2",
    "Wrist2": "Wrist2",
}

CLASS_ORDER = [
    "Bard",
    "Beastlord",
    "Berserker",
    "Cleric",
    "Druid",
    "Enchanter",
    "Magician",
    "Monk",
    "Necromancer",
    "Paladin",
    "Ranger",
    "Rogue",
    "Shadow Knight",
    "Shaman",
    "Warrior",
    "Wizard",
]


def lua_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def item_line(name: str, iid) -> str:
    name = str(name).strip()
    if iid is None or iid == "" or iid == "-":
        return name
    return f"{name}/{int(iid)}"


def parse_classes(cell) -> list[str]:
    if cell is None:
        return []
    text = str(cell).strip()
    if not text:
        return []
    upper = text.upper()
    if upper == "ALL":
        return list(ABBREV_TO_CLASS.values())
    if upper == "ALL MELEE":
        return [ABBREV_TO_CLASS[a] for a in ("WAR", "MNK", "ROG", "BER")]
    if upper == "ALL HYBRID":
        return [ABBREV_TO_CLASS[a] for a in ("SHD", "PAL", "BRD", "RNG", "BST")]
    if upper == "ALL CASTER":
        return [ABBREV_TO_CLASS[a] for a in ("WIZ", "MAG", "ENC", "NEC")]
    if upper == "ALL PRIEST":
        return [ABBREV_TO_CLASS[a] for a in ("CLR", "SHM", "DRU")]
    out = []
    for part in re.split(r"[/|,]+", text):
        part = part.strip().upper()
        if part in ABBREV_TO_CLASS:
            out.append(ABBREV_TO_CLASS[part])
    return out


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # --- Foci ID lookup ---
    foci_ids: dict[str, int] = {}
    ws = wb["Slot 12 Foci Augs"]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        iid = ws.cell(r, 2).value
        if name and iid:
            foci_ids[str(name).strip()] = int(iid)
            short = str(name).replace("Cryptic Clutch of ", "").strip()
            foci_ids[short] = int(iid)

    # --- Visible armor ---
    classes: dict[str, dict] = {c: {} for c in CLASS_ORDER}
    ws = wb["Visible Icons"]
    for r in range(2, ws.max_row + 1):
        abbr = ws.cell(r, 2).value
        slot_raw = ws.cell(r, 3).value
        iid = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if not abbr or not name:
            continue
        cls = ABBREV_TO_CLASS.get(str(abbr).strip().upper())
        slot = SLOT_MAP.get(str(slot_raw).strip())
        if not cls or not slot:
            continue
        classes[cls][slot] = item_line(name, iid)

    # --- Non-visible gear ---
    ear_lists: dict[str, list] = defaultdict(list)
    ring_lists: dict[str, list] = defaultdict(list)
    clicky_lists: dict[str, list] = defaultdict(list)
    ws = wb["Non-Visible"]
    for r in range(2, ws.max_row + 1):
        slot_raw = ws.cell(r, 1).value
        iid = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        classes_cell = ws.cell(r, 5).value
        if not slot_raw or not name:
            continue
        slot_raw = str(slot_raw).strip().strip('"')
        if slot_raw in ("Legendary Material", "Legendary") or name.startswith("Primary/"):
            continue
        if slot_raw in ("Range", "Charm", "Epic"):
            continue  # handled specially
        target_classes = parse_classes(classes_cell)
        if not target_classes:
            continue
        # Minion Materiel missing ID
        if "Minion" in str(name) and (iid is None or iid == ""):
            iid = 33110
        line = item_line(name, iid)
        if slot_raw == "Ear":
            for c in target_classes:
                ear_lists[c].append(line)
        elif slot_raw == "Ring":
            for c in target_classes:
                ring_lists[c].append(line)
        elif slot_raw == "Clicky":
            if "Minion" in str(name):
                continue  # not tracked on DoN BiS clicky rows
            for c in target_classes:
                clicky_lists[c].append(line)
        else:
            slot = {
                "Face": "Face",
                "Neck": "Neck",
                "Back": "Back",
                "Shoulder": "Shoulder",
                "Waist": "Waist",
            }.get(slot_raw)
            if not slot:
                continue
            for c in target_classes:
                classes[c][slot] = line

    for c, ears in ear_lists.items():
        if len(ears) >= 1:
            classes[c]["Ear1"] = ears[0]
        if len(ears) >= 2:
            classes[c]["Ear2"] = ears[1]
    for c, rings in ring_lists.items():
        if len(rings) >= 1:
            classes[c]["Finger1"] = rings[0]
        if len(rings) >= 2:
            classes[c]["Finger2"] = rings[1]
    for c, clicks in clicky_lists.items():
        for i, line in enumerate(clicks, start=1):
            key = "Clicky" if i == 1 else f"Clicky{i}"
            classes[c][key] = line

    # --- Range ---
    for c, (name, iid) in RANGE_BY_CLASS.items():
        classes[c]["Ranged"] = item_line(name, iid)

    # --- Legendary ---
    ws = wb["Class WeaponShield Components"]
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, 1).value
        name = ws.cell(r, 3).value
        abbr = ws.cell(r, 4).value
        if not name or not abbr or str(abbr).upper() == "ALL":
            continue
        if "Materium" in str(name):
            continue
        cls = ABBREV_TO_CLASS.get(str(abbr).strip().upper())
        if cls:
            classes[cls]["Legendary"] = item_line(name, iid)

    # --- Glyphs (item-only) ---
    glyph_by_template: dict[int, tuple[str, str, int]] = {}
    glyphs_by_class: dict[str, list] = defaultdict(list)
    ws = wb["Spell Clicky Components"]
    for r in range(2, ws.max_row + 1):
        gclass = ws.cell(r, 1).value
        spell = ws.cell(r, 2).value
        tmpl_id = ws.cell(r, 4).value
        clicky_name = ws.cell(r, 6).value
        clicky_id = ws.cell(r, 7).value
        if not gclass or not clicky_name:
            continue
        cls = GLYPH_CLASS_ALIASES.get(str(gclass).strip())
        if not cls:
            print("WARN unknown glyph class", gclass)
            continue
        line = item_line(clicky_name, clicky_id)
        glyphs_by_class[cls].append(line)
        if tmpl_id:
            glyph_by_template[int(tmpl_id)] = (cls, str(spell), int(clicky_id) if clicky_id else 0)

    for c, glyphs in glyphs_by_class.items():
        for i, line in enumerate(glyphs, start=1):
            classes[c][f"Glyph{i}"] = line

    # --- Packs + spell names ---
    pack_spells: dict[str, dict[str, list[str]]] = defaultdict(dict)
    announce_names: set[str] = set()

    def pack_spell_names(pack_name: str, template_ids: list[int]) -> list[str]:
        spells = []
        for tid in template_ids:
            if tid in glyph_by_template:
                spells.append(glyph_by_template[tid][1])
        if spells:
            return spells
        # Derive from pack item name
        m = re.match(r"^(?:Spell|Tome) Pack:\s*(.+)$", pack_name.strip())
        if m:
            return [m.group(1).strip()]
        return []

    ws = wb["Spell Packs"]
    pack_counts: dict[str, int] = defaultdict(int)
    for r in range(2, ws.max_row + 1):
        abbr = ws.cell(r, 1).value
        iid = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        if not abbr or not name:
            continue
        cls = ABBREV_TO_CLASS.get(str(abbr).strip().upper())
        if not cls:
            continue
        tmpl_ids = []
        for col in (7, 9, 11):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                tmpl_ids.append(int(v))
        pack_counts[cls] += 1
        n = pack_counts[cls]
        key = f"Pack{n}"
        line = item_line(name, iid)
        spells = pack_spell_names(str(name), tmpl_ids)
        if spells:
            # Encode as Lua table form for catalog builder
            classes[cls][key] = {"item": line, "spells": spells}
            pack_spells[cls][key] = spells
        else:
            classes[cls][key] = line
        announce_names.add(str(name).strip())

    # --- Foci per class ---
    for c, shorts in FOCI_BY_CLASS.items():
        for i, short in enumerate(shorts, start=1):
            full = f"Cryptic Clutch of {short}"
            iid = foci_ids.get(full) or foci_ids.get(short)
            classes[c][f"Aug{i}"] = item_line(full, iid)
            announce_names.add(full)

    # Collect announce names from all class items
    def collect_name(val):
        if isinstance(val, dict):
            val = val.get("item", "")
        name = str(val).split("/")[0].strip()
        if name:
            announce_names.add(name)

    for c, bucket in classes.items():
        for v in bucket.values():
            collect_name(v)

    # Template shared
    template = {
        "CharmExtreme": "Lustrous Gem of Unyielding Avarice/55054",
        "CharmSafe": "Faded Gem of Restrained Avarice/55055",
        "EpicUpgrade": "Scales of the Lava Dragon/66904",
        "Misc1": "Dark Reign Elite Satchel/66731",
        "Misc2": "Dark Reign Initiate Satchel/66732",
        "Misc3": "Ancient Draconic Lockbox I/66733",
        "Misc4": "A Strange Compass/66721",
        "Misc5": "Radiant Crystal Cache/66902",
        "Misc6": "Duality of Desire/66903",
        "Materium1": "Primary Materium of Legends/62535",
        "Materium2": "Secondary Materium of Legends/62536",
        "Materium3": "Tertiary Materium of Legends/62537",
    }
    for v in template.values():
        collect_name(v)

    # Vacant vessels → announce only
    ws = wb["Slot 12 Foci Augs"]
    for r in range(2, ws.max_row + 1):
        vessel = ws.cell(r, 8).value
        if vessel:
            announce_names.add(str(vessel).strip())

    # Max slot counts for Main.Slots
    max_aug = max(len(FOCI_BY_CLASS[c]) for c in CLASS_ORDER)
    max_pack = max(pack_counts.values()) if pack_counts else 1
    max_glyph = max((len(glyphs_by_class[c]) for c in CLASS_ORDER), default=1)
    max_clicky = max((len(clicky_lists[c]) for c in CLASS_ORDER), default=1)

    clicky_slots = ["Clicky"] + [f"Clicky{i}" for i in range(2, max_clicky + 1)]
    pack_slots = [f"Pack{i}" for i in range(1, max_pack + 1)]
    glyph_slots = [f"Glyph{i}" for i in range(1, max_glyph + 1)]
    aug_slots = [f"Aug{i}" for i in range(1, max_aug + 1)]

    # --- Emit Lua block ---
    lines = []
    lines.append("\t['don'] = {")

    def emit_value(val, indent: str) -> str:
        if isinstance(val, dict):
            item = val["item"]
            spells = val.get("spells") or []
            spell_lits = ", ".join(f"'{lua_escape(s)}'" for s in spells)
            return (
                f"{{\n"
                f"{indent}\titem = '{lua_escape(item)}',\n"
                f"{indent}\tspells = {{{spell_lits}}},\n"
                f"{indent}}}"
            )
        return f"'{lua_escape(val)}'"

    for c in CLASS_ORDER:
        bucket = classes[c]
        lines.append(f"\t\t['{c}'] = {{")
        # Stable slot order
        preferred = (
            ["Head", "Chest", "Arms", "Wrist1", "Hands", "Legs", "Feet", "Wrist2"]
            + ["Ear1", "Ear2", "Face", "Neck", "Back", "Shoulder", "Waist", "Finger1", "Finger2"]
            + ["Ranged"]
            + clicky_slots
            + ["Legendary"]
            + pack_slots
            + glyph_slots
            + aug_slots
        )
        seen = set()
        for slot in preferred:
            if slot in bucket:
                lines.append(f"\t\t\t['{slot}'] = {emit_value(bucket[slot], chr(9)*3)},")
                seen.add(slot)
        for slot in sorted(bucket.keys()):
            if slot not in seen:
                lines.append(f"\t\t\t['{slot}'] = {emit_value(bucket[slot], chr(9)*3)},")
        lines.append("\t\t},")

    lines.append("\t\t['Main'] = {")
    lines.append("\t\t\t['Slots'] = {")
    lines.append(
        "\t\t\t\t{Name='Visibles', Slots={'Head','Chest','Arms','Wrist1','Hands','Legs','Feet','Wrist2',}},"
    )
    lines.append(
        "\t\t\t\t{Name='Non-Visibles', Slots={'Ear1','Ear2','Face','Neck','Back','Shoulder','Waist','Finger1','Finger2',}},"
    )
    lines.append(
        "\t\t\t\t{Name='Drakes', Slots={'Ranged','CharmExtreme','CharmSafe',}},"
    )
    clicky_lua = ",".join(f"'{s}'" for s in clicky_slots)
    lines.append(
        f"\t\t\t\t{{Name='Clickies', Slots={{{clicky_lua},'EpicUpgrade','Legendary',}}}},"
    )
    pack_lua = ",".join(f"'{s}'" for s in pack_slots)
    lines.append(f"\t\t\t\t{{Name='Packs', Slots={{{pack_lua},}}}},")
    glyph_lua = ",".join(f"'{s}'" for s in glyph_slots)
    lines.append(f"\t\t\t\t{{Name='Glyphs', Slots={{{glyph_lua},}}}},")
    aug_lua = ",".join(f"'{s}'" for s in aug_slots)
    lines.append(f"\t\t\t\t{{Name='Foci', Slots={{{aug_lua},}}}},")
    lines.append(
        "\t\t\t\t{Name='Misc', "
        "Slots={'Misc1','Misc2','Misc3','Misc4','Misc5','Misc6',"
        "'Materium1','Materium2','Materium3',}},"
    )
    lines.append("\t\t\t},")
    lines.append("\t\t},")
    lines.append("\t\t['Template'] = {")
    for k, v in template.items():
        lines.append(f"\t\t\t['{k}'] = '{lua_escape(v)}',")
    lines.append("\t\t},")
    lines.append("\t},")

    OUT_BLOCK.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_BLOCK}")

    # Announce names sorted
    sorted_names = sorted(announce_names, key=lambda s: s.lower())
    OUT_ANNOUNCE.write_text("\n".join(sorted_names) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_ANNOUNCE} ({len(sorted_names)} names)")

    # Spot checks
    for check in ("Warrior", "Ranger", "Shadow Knight", "Cleric"):
        b = classes[check]
        print(f"\n== {check} ==")
        for k in ("Head", "Ear1", "Ear2", "Ranged", "Clicky", "Pack1", "Glyph1", "Aug1", "Legendary"):
            if k in b:
                print(f"  {k}: {b[k]}")


if __name__ == "__main__":
    main()
