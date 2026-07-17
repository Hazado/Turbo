import openpyxl

wb = openpyxl.load_workbook(r"d:\Desktop\Turbo Wednesday 17th\DoN Misc..xlsx", data_only=True)
for name in [
    "Spell Clicky Components",
    "Spell Packs",
    "Non-Visible",
    "Visible Icons",
    "Slot 12 Foci Augs",
    "Misc. Items",
]:
    ws = wb[name]
    print("===", name, "rows", ws.max_row, "cols", ws.max_column)
    for r in range(1, min(8, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 14) + 1)]
        print(r, vals)
    print()
